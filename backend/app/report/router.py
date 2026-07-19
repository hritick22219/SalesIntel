from datetime import date
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db
from models import SalesRecord, User
from auth.dependencies import get_current_user

router = APIRouter(tags=["Reports"])

# --- Helper Query Function ---
async def get_filtered_records(
    db: AsyncSession,
    company_id: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    product: Optional[str] = None,
    customer: Optional[str] = None,
):
    query = select(SalesRecord).filter(SalesRecord.company_id == company_id)

    if start_date:
        query = query.filter(SalesRecord.date >= start_date)
    if end_date:
        query = query.filter(SalesRecord.date <= end_date)
    if product:
        query = query.filter(SalesRecord.product.ilike(f"%{product}%"))
    if customer:
        query = query.filter(SalesRecord.customer.ilike(f"%{customer}%"))

    query = query.order_by(SalesRecord.date.desc())
    res = await db.execute(query)
    records = res.scalars().all()

    # Format into list of dicts
    data = []
    for r in records:
        data.append({
            "Date": r.date.isoformat() if isinstance(r.date, date) else r.date,
            "Product": r.product,
            "Customer": r.customer,
            "Category": r.category or "",
            "Revenue": float(r.revenue),
            "Profit": float(r.profit)
        })

    df = pd.DataFrame(data)
    if df.empty:
        # Create schema for empty reports
        df = pd.DataFrame(columns=["Date", "Product", "Customer", "Category", "Revenue", "Profit"])
    return df

# --- Routes ---

@router.get("/csv")
async def export_csv(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    product: Optional[str] = None,
    customer: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    df = await get_filtered_records(db, current_user.company_id, start_date, end_date, product, customer)

    buffer = io.StringIO()
    df.to_csv(buffer, index=False)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sales_report.csv"}
    )

@router.get("/excel")
async def export_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    product: Optional[str] = None,
    customer: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    df = await get_filtered_records(db, current_user.company_id, start_date, end_date, product, customer)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sales Records")

        workbook = writer.book
        worksheet = writer.sheets["Sales Records"]

        # --- Premium Styling Configuration ---
        # Fonts
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10)
        
        # Colors & Fills
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Sleek Dark Blue
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # Light Gray
        
        # Alignments
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Style Headers
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Style Content rows
        currency_format = "$#,##0.00"
        for row_idx in range(2, len(df) + 2):
            # Center Date
            worksheet.cell(row=row_idx, column=1).alignment = center_align
            # Left align Product, Customer, Category
            worksheet.cell(row=row_idx, column=2).alignment = left_align
            worksheet.cell(row=row_idx, column=3).alignment = left_align
            worksheet.cell(row=row_idx, column=4).alignment = left_align
            
            # Format Revenue (Column 5)
            rev_cell = worksheet.cell(row=row_idx, column=5)
            rev_cell.number_format = currency_format
            rev_cell.alignment = right_align
            
            # Format Profit (Column 6)
            prof_cell = worksheet.cell(row=row_idx, column=6)
            prof_cell.number_format = currency_format
            prof_cell.alignment = right_align

            # Zebra stripe rows for readability
            if row_idx % 2 == 0:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = zebra_fill

            # Apply font to all cells in the row
            for col_idx in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).font = regular_font

        # Auto-adjust column widths based on maximum text length
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                # Estimate currency formatting spacing
                if isinstance(val, float):
                    val_len = len(f"${val:,.2f}")
                else:
                    val_len = len(str(val or ''))
                if val_len > max_len:
                    max_len = val_len
            # Apply padding
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 11)

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"}
    )
